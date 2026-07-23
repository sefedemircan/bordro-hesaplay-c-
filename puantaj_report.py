"""Meyer puantaj verisini sade, aylık bir puantaj raporuna dönüştürür."""

from __future__ import annotations

import hashlib
import io
import math
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime, time, timedelta
from typing import BinaryIO

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


WEEKLY_MAX_HOURS = 45.0
FULL_DAY_HOURS = 9.0

# Yasal mazeret / izin / resmi tatil: çalışılmış sayılır, hafta tatili kesilmez.
SUNDAY_PROTECTING_STATUSES = frozenset({
    "RAPOR",
    "YILLIK_IZIN",
    "UCRETLI_IZIN",
    "RESMI_TATIL",
    "YARIM_GUN",
    "UZAKTAN",
    "CALISMA",
    "SERBEST_ZAMAN",
    "HARIC",
})

STATUS_LABELS = {
    "CALISMA": "Çalışma",
    "YILLIK_IZIN": "Yıllık İzin",
    "UCRETLI_IZIN": "Ücretli İzin",
    "RAPOR": "Rapor",
    "UCRETSIZ_IZIN": "Ücretsiz İzin",
    "DEVAMSIZ": "Devamsızlık",
    "HAFTA_TATILI": "Hafta Tatili",
    "UCRETSIZ_HAFTA_TATILI": "Ücretsiz Hafta Tatili",
    "RESMI_TATIL": "Resmi Tatil",
    "SERBEST_ZAMAN": "Serbest Zaman",
    "HARIC": "Sayılmayan Gün",
    "UZAKTAN": "Uzaktan Çalışma",
    "YARIM_GUN": "Yarım Gün",
}

STATUS_CODES = {
    "YILLIK_IZIN": "Y",
    "UCRETLI_IZIN": "Ü",
    "RAPOR": "R",
    "UCRETSIZ_IZIN": "Z",
    "DEVAMSIZ": "M",
    "HAFTA_TATILI": "T",
    "UCRETSIZ_HAFTA_TATILI": "C",
    "RESMI_TATIL": "B",
    "SERBEST_ZAMAN": "A3",
    "HARIC": "X",
    "UZAKTAN": "U",
    "YARIM_GUN": "V",
}

CODE_LEGEND = [
    ("Saat", "Fiili çalışma (NM + FM)"),
    ("N", "Normal çalışma (9,00)"),
    ("U", "Uzaktan çalışma (9,00)"),
    ("R", "Raporlu (7,50)"),
    ("W", "Tatilde raporlu (7,50)"),
    ("Y", "Yıllık izin (7,50)"),
    ("Ü", "Ücretli izin (7,50)"),
    ("Z", "Ücretsiz izin / mazeretsiz nedeniyle kesilen hafta tatili (7,50)"),
    ("M", "Mazeretsiz (7,50)"),
    ("T", "Hafta tatili / resmi tatil (7,50)"),
    ("C", "Ücretsiz hafta tatili (7,50)"),
    ("B", "Resmi tatil hafta içi (7,50)"),
    ("D", "Dini bayram (7,50)"),
    ("K", "Yarım gün resmi tatil (3,75)"),
    ("V", "Yarım gün (3,75)"),
    ("X", "Sayılmayan gün (9,00)"),
    ("A1", "Serbest zaman maktu tam gün (1)"),
    ("A2", "Serbest zaman maktu yarım gün (0,5)"),
    ("A3", "Serbest zaman saatlik tam gün / cumartesi mesaisiz (0)"),
    ("A4", "Serbest zaman saatlik yarım gün (3,75)"),
]

EXCEL_CODE_COLORS = {
    "T": "D9EAD3",
    "Z": "F4CCCC",
    "C": "F4CCCC",
    "Y": "D9EAF7",
    "Ü": "CFE2F3",
    "R": "FFF2CC",
    "W": "FFF2CC",
    "M": "FCE5CD",
    "A3": "EAD1DC",
    "A1": "EAD1DC",
    "A2": "EAD1DC",
    "A4": "EAD1DC",
    "B": "D0E0E3",
    "D": "D0E0E3",
    "K": "D0E0E3",
    "V": "D9D2E9",
    "X": "EFEFEF",
    "U": "D9EAD3",
}


@dataclass
class ReportResult:
    daily: pd.DataFrame
    monthly: pd.DataFrame
    summary: pd.DataFrame
    weekly: pd.DataFrame
    quality: pd.DataFrame
    period_start: pd.Timestamp
    period_end: pd.Timestamp


def _normalized_text(value: object) -> str:
    text = str(value).strip().replace("ı", "i").replace("İ", "I")
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch)).upper()


def _column_key(value: object) -> str:
    return re.sub(r"[^A-Z0-9]", "", _normalized_text(value))


CANONICAL_COLUMNS = {
    "SICILNO": "sicilno",
    "AD": "Ad",
    "SOYAD": "Soyad",
    "FIRMA": "Firma",
    "BOLUM": "Bölüm",
    "POZISYON": "Pozisyon",
    "GOREV": "Görev",
    "YAKA": "Yaka",
    "MESAITARIH": "mesaitarih",
    "GIRIS": "Giriş",
    "CIKIS": "Çıkış",
    "MS": "MS",
    "NM": "NM",
    "FM": "FM",
    "IZS": "IZS",
    "YIZS": "YIZS",
    "SGKIZS": "SGKIZS",
    "UCZIZS": "UCZIZS",
    "RM": "RM",
    "EM": "EM",
    "IZINACIKLAMA": "İzin Açıklama",
    "MESAIACIKLAMA": "Mesai Açıklama",
}

SAKRA_DAY_BLOCKS = (25, 44, 63, 82, 101, 120)  # Y, AR, BK, CD, CW, DP
SAKRA_NORMAL_COLUMNS = (37, 56, 75, 94, 113, 132)  # AK, BD, BW, CP, DI, EB
SAKRA_OVERTIME_COLUMNS = (39, 58, 77, 96, 115, 134)  # AM, BF, BY, CR, DK, ED


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename = {}
    occupied = set(df.columns)
    for column in df.columns:
        canonical = CANONICAL_COLUMNS.get(_column_key(column))
        if canonical and (canonical == column or canonical not in occupied):
            rename[column] = canonical
    return df.rename(columns=rename).copy()


def _is_sakra_sheet(ws) -> bool:
    return (
        _column_key(ws.cell(2, 3).value) == "ADISOYADI"
        and _column_key(ws.cell(2, 13).value) == "YIL"
        and _column_key(ws.cell(2, 14).value) == "AY"
    )


def _numeric_hours(value: object) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)) and not pd.isna(value):
        return float(value) if float(value) > 0 else None
    return None


def _sakra_code_fields(value: object) -> dict[str, object]:
    code = str(value).strip() if value is not None else ""
    normalized = _normalized_text(code)
    fields: dict[str, object] = {"Kaynak Kod": code}
    if normalized in {"R", "W"}:
        fields["SGKIZS"] = "07:30:00"
    elif normalized in {"Y", "S"}:
        fields["YIZS"] = "07:30:00"
    elif code in {"Ü", "ü"} or normalized == "I":
        fields["IZS"] = "07:30:00"
    elif normalized in {"M", "E"}:
        fields["EM"] = "09:00:00"
    elif normalized == "Z":
        fields["UCZIZS"] = "07:30:00"
    elif normalized == "C":
        fields["UCZIZS"] = "07:30:00"
    elif normalized == "V":
        fields["IZS"] = "03:45:00"
    return fields


def _source_status_from_code(raw_code: object) -> str | None:
    """Kaynak harf kodunu durum etiketine çevirir (görsel + eski Sakra)."""
    code = str(raw_code).strip() if raw_code is not None else ""
    if not code:
        return None
    if code in {"Ü", "ü"}:
        return "UCRETLI_IZIN"
    normalized = _normalized_text(code)
    return {
        "R": "RAPOR",
        "W": "RAPOR",
        "Y": "YILLIK_IZIN",
        "S": "YILLIK_IZIN",  # eski Sakra
        "I": "UCRETLI_IZIN",  # eski Sakra İ/I
        "Z": "UCRETSIZ_IZIN",
        "M": "DEVAMSIZ",
        "E": "DEVAMSIZ",  # eski Sakra
        "T": "HAFTA_TATILI",
        "H": "HAFTA_TATILI",  # eski Sakra
        "C": "UCRETSIZ_HAFTA_TATILI",
        "X": "HARIC",
        "B": "RESMI_TATIL",
        "D": "RESMI_TATIL",
        "K": "RESMI_TATIL",
        "U": "UZAKTAN",
        "N": "CALISMA",
        "V": "YARIM_GUN",
        "A1": "SERBEST_ZAMAN",
        "A2": "SERBEST_ZAMAN",
        "A3": "SERBEST_ZAMAN",
        "A4": "SERBEST_ZAMAN",
    }.get(normalized)


def _classify_row(row: pd.Series) -> str:
    source_status = _source_status_from_code(row.get("Kaynak Kod", ""))
    if source_status:
        return source_status
    description = _normalized_text(row.get("İzin Açıklama", ""))
    if row["SGKIZS_h"] > 0 or row["RM_h"] > 0 or "RAPOR" in description or "SGK" in description:
        return "RAPOR"
    if row["YIZS_h"] > 0 or "YILLIK" in description:
        return "YILLIK_IZIN"
    if row["UCZIZS_h"] > 0 or "UCRETSIZ" in description:
        return "UCRETSIZ_IZIN"
    if row["IZS_h"] > 0 or ("IZIN" in description and "#__#" not in description):
        return "UCRETLI_IZIN"
    if row["NM_h"] > 0 or row["FM_h"] > 0:
        return "CALISMA"
    if row["is_weekend"]:
        return "HAFTA_TATILI"
    # Meyer'deki EM eksik mesai alanıdır; tek başına izin kanıtı değildir.
    return "DEVAMSIZ"


def _read_sakra_workbook(file_or_buffer: str | BinaryIO) -> pd.DataFrame | None:
    if hasattr(file_or_buffer, "seek"):
        file_or_buffer.seek(0)
    wb = load_workbook(file_or_buffer, data_only=True, read_only=False)
    try:
        ws = next((sheet for sheet in wb.worksheets if _is_sakra_sheet(sheet)), None)
        if ws is None:
            return None

        rows: list[dict[str, object]] = []
        for row_number in range(3, ws.max_row + 1):
            full_name = str(ws.cell(row_number, 3).value or "").strip()
            year_value = ws.cell(row_number, 13).value
            month_value = ws.cell(row_number, 14).value
            if not full_name or not isinstance(year_value, (int, float)) or not isinstance(month_value, (int, float)):
                continue
            year, month = int(year_value), int(month_value)
            if not (1900 <= year <= 2200 and 1 <= month <= 12):
                continue

            employee_number = ws.cell(row_number, 2).value
            if employee_number in (None, ""):
                digest = hashlib.sha1(_normalized_text(full_name).encode("utf-8")).hexdigest()[:8].upper()
                employee_number = f"SKR-{digest}"

            first_weekday = pd.Timestamp(year, month, 1).dayofweek
            days_in_month = int((pd.Timestamp(year, month, 1) + pd.offsets.MonthEnd()).day)
            identity = {
                "sicilno": str(employee_number),
                "Ad": full_name,
                "Soyad": "",
                "Firma": ws.cell(row_number, 5).value or "",
                "Bölüm": ws.cell(row_number, 7).value or "",
                "Pozisyon": ws.cell(row_number, 4).value or "",
                "Görev": ws.cell(row_number, 4).value or "",
                "Yaka": ws.cell(row_number, 8).value or "",
                "Hesaplanmış": True,
            }

            week_entries: dict[int, list[dict[str, object]]] = {idx: [] for idx in range(6)}
            for day in range(1, days_in_month + 1):
                slot = first_weekday + day - 1
                week_index, weekday = divmod(slot, 7)
                cell_value = ws.cell(row_number, SAKRA_DAY_BLOCKS[week_index] + weekday).value
                hours = _numeric_hours(cell_value)
                if hours is None and cell_value in (None, "", 0):
                    continue
                record = {
                    **identity,
                    "mesaitarih": pd.Timestamp(year, month, day),
                    "MS": 0, "NM": 0, "FM": 0, "IZS": 0, "YIZS": 0,
                    "SGKIZS": 0, "UCZIZS": 0, "RM": 0, "EM": 0,
                    "İzin Açıklama": "",
                }
                if hours is not None:
                    record["Toplam Çalışma"] = hours
                else:
                    record.update(_sakra_code_fields(cell_value))
                week_entries[week_index].append(record)

            # Cached haftalık sonuçları günlük toplamı değiştirmeden NM/FM olarak dağıt.
            for week_index, entries in week_entries.items():
                normal_remaining = _numeric_hours(ws.cell(row_number, SAKRA_NORMAL_COLUMNS[week_index]).value) or 0.0
                overtime_remaining = _numeric_hours(ws.cell(row_number, SAKRA_OVERTIME_COLUMNS[week_index]).value) or 0.0
                for record in entries:
                    total = float(record.pop("Toplam Çalışma", 0.0))
                    if total > 0:
                        normal = min(total, normal_remaining)
                        normal_remaining -= normal
                        remainder = total - normal
                        overtime = min(remainder, overtime_remaining)
                        overtime_remaining -= overtime
                        record["NM"] = normal + max(0.0, remainder - overtime)
                        record["FM"] = overtime
                    rows.append(record)

        if not rows:
            raise ValueError("Sakra sayfasında raporlanabilir personel-gün verisi bulunamadı.")
        return normalize_columns(pd.DataFrame(rows))
    finally:
        wb.close()
        if hasattr(file_or_buffer, "seek"):
            file_or_buffer.seek(0)


def read_puantaj_file(file_or_buffer: str | BinaryIO, filename: object | None = None) -> pd.DataFrame:
    raw_name = filename if filename not in (None, "") else getattr(file_or_buffer, "name", "")
    name = str(raw_name).lower()
    if name.endswith(".csv"):
        raw = file_or_buffer.read() if hasattr(file_or_buffer, "read") else open(file_or_buffer, "rb").read()
        if isinstance(raw, str):
            raw = raw.encode("utf-8")
        last_error = None
        for encoding in ("utf-8-sig", "cp1254", "utf-8", "latin-1"):
            try:
                return normalize_columns(pd.read_csv(io.BytesIO(raw), sep=";", encoding=encoding))
            except UnicodeDecodeError as exc:
                last_error = exc
        raise ValueError("CSV karakter kodlaması okunamadı.") from last_error

    if name.endswith(".xlsx"):
        sakra = _read_sakra_workbook(file_or_buffer)
        if sakra is not None:
            return sakra
    engine = "xlrd" if name.endswith(".xls") and not name.endswith(".xlsx") else "openpyxl"
    return normalize_columns(pd.read_excel(file_or_buffer, engine=engine))


def time_to_hours(value: object) -> float:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return 0.0
    if isinstance(value, (pd.Timedelta, timedelta)):
        return value.total_seconds() / 3600
    if isinstance(value, (pd.Timestamp, datetime, time)):
        return value.hour + value.minute / 60 + value.second / 3600
    if isinstance(value, (int, float)):
        # Excel saatleri günün kesri olarak saklar; 1'den büyük sayılar saat kabul edilir.
        return float(value) * 24 if 0 <= float(value) < 1 else float(value)

    text = str(value).strip()
    if not text or text in {"#__#", "-"}:
        return 0.0
    try:
        if "day" in text.lower():
            return pd.to_timedelta(text).total_seconds() / 3600
        parts = text.split(":")
        if len(parts) >= 2:
            return float(parts[0]) + float(parts[1]) / 60 + (float(parts[2]) / 3600 if len(parts) > 2 else 0)
        return float(text.replace(",", "."))
    except (TypeError, ValueError):
        return 0.0


def format_hours(value: float, zero: str = "00:00") -> str:
    if pd.isna(value) or abs(float(value)) < 1e-9:
        return zero
    sign = "-" if value < 0 else ""
    minutes = int(round(abs(float(value)) * 60))
    return f"{sign}{minutes // 60:02d}:{minutes % 60:02d}"


def _hours_series(df: pd.DataFrame, column: str) -> pd.Series:
    if column not in df.columns:
        return pd.Series(0.0, index=df.index, dtype=float)
    return df[column].map(time_to_hours).astype(float)


def _employee_id(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if re.fullmatch(r"\d+(?:\.0)?", text):
        return str(int(float(text))).zfill(5)
    return text


def prepare_daily(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = normalize_columns(df)
    required = {"sicilno", "Ad", "Soyad", "mesaitarih", "NM", "FM"}
    missing = sorted(required - set(df.columns))
    if missing:
        raise ValueError("Zorunlu sütunlar eksik: " + ", ".join(missing))

    quality_rows = []
    work = df.copy()
    work["Tarih"] = pd.to_datetime(work["mesaitarih"], dayfirst=True, errors="coerce")
    invalid_dates = int(work["Tarih"].isna().sum())
    if invalid_dates:
        quality_rows.append({"Seviye": "Hata", "Kontrol": "Geçersiz tarih", "Kayıt": invalid_dates, "Açıklama": "Bu satırlar rapora alınmadı."})
    work = work.dropna(subset=["Tarih"]).copy()
    work["sicilno"] = work["sicilno"].map(_employee_id)
    missing_ids = int(work["sicilno"].eq("").sum())
    if missing_ids:
        quality_rows.append({"Seviye": "Hata", "Kontrol": "Boş sicil no", "Kayıt": missing_ids, "Açıklama": "Bu satırlar rapora alınmadı."})
        work = work[work["sicilno"].ne("")].copy()

    duplicate_mask = work.duplicated(["sicilno", "Tarih"], keep=False)
    duplicates = int(duplicate_mask.sum())
    if duplicates:
        quality_rows.append({"Seviye": "Uyarı", "Kontrol": "Mükerrer kişi-gün", "Kayıt": duplicates, "Açıklama": "Süreler birleştirilmeden önce ilk kayıt esas alındı."})
        work = work.drop_duplicates(["sicilno", "Tarih"], keep="first")

    for column in ("MS", "NM", "FM", "IZS", "YIZS", "SGKIZS", "UCZIZS", "RM", "EM"):
        work[f"{column}_h"] = _hours_series(work, column)
    for column in ("Firma", "Bölüm", "Pozisyon", "Görev", "Yaka", "Giriş", "Çıkış", "İzin Açıklama", "Kaynak Kod"):
        if column not in work.columns:
            work[column] = ""
    if "Hesaplanmış" not in work.columns:
        work["Hesaplanmış"] = False

    work["Personel"] = (work["Ad"].fillna("").astype(str).str.strip() + " " + work["Soyad"].fillna("").astype(str).str.strip()).str.strip()
    work["day_of_week"] = work["Tarih"].dt.dayofweek
    work["is_weekend"] = work["day_of_week"] >= 5
    iso = work["Tarih"].dt.isocalendar()
    work["iso_year"] = iso.year.astype(int)
    work["iso_week"] = iso.week.astype(int)
    work["Durum"] = work.apply(_classify_row, axis=1)
    work["NM Güncel_h"] = work["NM_h"]
    work["FM Güncel_h"] = work["FM_h"]
    work["FM→NM_h"] = 0.0
    work["Pazar Durumu"] = ""

    weekly_rows = []
    group_keys = ["sicilno", "iso_year", "iso_week"]
    for (sicilno, iso_year, iso_week), indices in work.groupby(group_keys, sort=True).groups.items():
        idx = list(indices)
        week = work.loc[idx]
        weekday_idx = week.index[week["day_of_week"] < 5].tolist()
        weekend_idx = week.index[week["day_of_week"] >= 5].tolist()
        transferred = 0.0
        if not bool(week["Hesaplanmış"].fillna(False).any()):
            missing_nm = max(0.0, WEEKLY_MAX_HOURS - work.loc[weekday_idx, "NM Güncel_h"].sum())
            for row_idx in weekend_idx:
                amount = min(missing_nm, work.at[row_idx, "FM Güncel_h"])
                if amount > 0:
                    work.at[row_idx, "FM Güncel_h"] -= amount
                    work.at[row_idx, "NM Güncel_h"] += amount
                    work.at[row_idx, "FM→NM_h"] += amount
                    transferred += amount
                    missing_nm -= amount

        sunday_idx = week.index[week["day_of_week"] == 6].tolist()
        source_burned = bool(work.loc[sunday_idx, "Durum"].eq("UCRETSIZ_HAFTA_TATILI").any())
        # Yalnızca mazeretsiz/devamsızlık keser; rapor, ücretli izin, resmi tatil vb. çalışılmış sayılır.
        weekday_slice = work.loc[weekday_idx]
        unprotected = weekday_slice["Durum"].eq("DEVAMSIZ") & (
            weekday_slice["EM_h"].fillna(0) >= FULL_DAY_HOURS - 0.01
        )
        # Korunan durumlar açıkça dışlanır (ileride sınıflama değişirse güvence).
        protected = weekday_slice["Durum"].isin(SUNDAY_PROTECTING_STATUSES)
        full_absence = source_burned or bool((unprotected & ~protected).any())
        work.loc[sunday_idx, "Pazar Durumu"] = "Kesildi" if full_absence else "Hak Edildi"
        if full_absence:
            work.loc[sunday_idx, "Hafta Tatili Kodu"] = "Z"

        week_start = pd.Timestamp.fromisocalendar(int(iso_year), int(iso_week), 1)
        week_end = week_start + pd.Timedelta(days=6)
        person = week.iloc[0]
        weekly_rows.append({
            "Sicil No": sicilno,
            "Personel": person["Personel"],
            "Bölüm": person["Bölüm"],
            "Hafta": f"{iso_year}-H{iso_week:02d}",
            "Hafta Aralığı": f"{week_start:%d.%m.%Y} – {week_end:%d.%m.%Y}",
            "Hafta İçi NM": work.loc[weekday_idx, "NM Güncel_h"].sum(),
            "Hafta Sonu Ham FM": week.loc[weekend_idx, "FM_h"].sum(),
            "FM→NM Aktarım": transferred,
            "Toplam NM": work.loc[idx, "NM Güncel_h"].sum(),
            "Kalan FM": work.loc[idx, "FM Güncel_h"].sum(),
            "Pazar Durumu": "Kesildi" if full_absence else "Hak Edildi",
        })

    work["Durum Açıklaması"] = work["Durum"].map(STATUS_LABELS)
    work["Kod"] = work["Durum"].map(STATUS_CODES).astype(object)
    working = work["Durum"].isin({"CALISMA", "UZAKTAN"})
    work.loc[working, "Kod"] = (work.loc[working, "NM_h"] + work.loc[working, "FM_h"]).map(lambda h: round(h, 2))
    # Uzaktan çalışma saati yoksa U harfi kalsın.
    remote_no_hours = work["Durum"].eq("UZAKTAN") & (work["NM_h"] + work["FM_h"] <= 0.01)
    work.loc[remote_no_hours, "Kod"] = "U"
    source_code_mask = work["Kaynak Kod"].fillna("").astype(str).str.strip().ne("") & ~working
    work.loc[source_code_mask, "Kod"] = work.loc[source_code_mask, "Kaynak Kod"]
    burned_sunday = (work["day_of_week"] == 6) & work["Pazar Durumu"].eq("Kesildi") & work["Durum"].eq("HAFTA_TATILI")
    work.loc[burned_sunday, "Kod"] = "Z"

    # Cumartesi mesai yoksa A3 (serbest zaman saatlik tam gün).
    saturday_no_mesai = (
        work["day_of_week"].eq(5)
        & work["Durum"].eq("HAFTA_TATILI")
        & ((work["NM_h"] + work["FM_h"]) <= 0.01)
    )
    work.loc[saturday_no_mesai, "Durum"] = "SERBEST_ZAMAN"
    work.loc[saturday_no_mesai, "Durum Açıklaması"] = STATUS_LABELS["SERBEST_ZAMAN"]
    work.loc[saturday_no_mesai, "Kod"] = "A3"

    if not quality_rows:
        quality_rows.append({"Seviye": "Bilgi", "Kontrol": "Dosya yapısı", "Kayıt": len(work), "Açıklama": "Zorunlu alanlar ve kişi-gün anahtarı uygun."})
    quality = pd.DataFrame(quality_rows)
    weekly = pd.DataFrame(weekly_rows)
    return work.sort_values(["Personel", "Tarih"]).reset_index(drop=True), weekly, quality


def _period_mask(daily: pd.DataFrame, year: int, month: int) -> pd.Series:
    return daily["Tarih"].dt.year.eq(year) & daily["Tarih"].dt.month.eq(month)


def build_report(df: pd.DataFrame, year: int | None = None, month: int | None = None) -> ReportResult:
    daily_all, weekly_all, quality = prepare_daily(df)
    if daily_all.empty:
        raise ValueError("Raporlanabilir kayıt bulunamadı.")
    if year is None or month is None:
        first_period = daily_all["Tarih"].dt.to_period("M").min()
        year, month = int(first_period.year), int(first_period.month)
    daily = daily_all[_period_mask(daily_all, year, month)].copy()
    if daily.empty:
        raise ValueError(f"{month:02d}.{year} döneminde kayıt bulunamadı.")

    period_start = pd.Timestamp(year=year, month=month, day=1)
    period_end = period_start + pd.offsets.MonthEnd(1)
    weekly_keys = set(zip(daily["sicilno"], daily["iso_year"], daily["iso_week"]))
    weekly = weekly_all[
        weekly_all.apply(lambda r: (r["Sicil No"], int(r["Hafta"].split("-H")[0]), int(r["Hafta"].split("-H")[1])) in weekly_keys, axis=1)
    ].copy()

    identity = ["sicilno", "Personel", "Firma", "Bölüm", "Pozisyon", "Görev", "Yaka"]
    employee_info = daily.sort_values("Tarih").groupby("sicilno", as_index=False)[identity[1:]].first()
    daily["Ay Günü"] = daily["Tarih"].dt.day
    grid = daily.pivot(index="sicilno", columns="Ay Günü", values="Kod")
    grid = grid.reindex(columns=range(1, period_end.day + 1))
    day_names = ("Pt", "Sa", "Ça", "Pe", "Cu", "Ct", "Pz")
    grid.columns = [f"{day:02d} {day_names[pd.Timestamp(year, month, day).dayofweek]}" for day in grid.columns]
    monthly = employee_info.merge(grid.reset_index(), on="sicilno", how="left").rename(columns={"sicilno": "Sicil No"})

    records = []
    for sicilno, group in daily.groupby("sicilno", sort=True):
        person = group.iloc[0]
        status_counts = group["Durum"].value_counts()
        records.append({
            "Sicil No": sicilno,
            "Personel": person["Personel"],
            "Firma": person["Firma"],
            "Bölüm": person["Bölüm"],
            "Pozisyon": person["Pozisyon"],
            "Çalışma Günü": int(status_counts.get("CALISMA", 0)),
            "Normal Çalışma": group["NM Güncel_h"].sum(),
            "Fazla Mesai": group["FM Güncel_h"].sum(),
            "FM→NM Aktarım": group["FM→NM_h"].sum(),
            "Yıllık İzin (gün)": int(status_counts.get("YILLIK_IZIN", 0)),
            "Ücretli İzin (gün)": int(status_counts.get("UCRETLI_IZIN", 0)),
            "Rapor (gün)": int(status_counts.get("RAPOR", 0)),
            "Ücretsiz İzin (gün)": int(status_counts.get("UCRETSIZ_IZIN", 0)),
            "Devamsızlık (gün)": int(status_counts.get("DEVAMSIZ", 0)),
            "Hafta Tatili (gün)": int(status_counts.get("HAFTA_TATILI", 0)),
            "Ücretsiz Hafta Tatili (gün)": int(status_counts.get("UCRETSIZ_HAFTA_TATILI", 0)),
            "Resmi Tatil (gün)": int(status_counts.get("RESMI_TATIL", 0)),
            "Serbest Zaman (gün)": int(status_counts.get("SERBEST_ZAMAN", 0)),
            "Uzaktan Çalışma (gün)": int(status_counts.get("UZAKTAN", 0)),
            "Yarım Gün": int(status_counts.get("YARIM_GUN", 0)),
            "Sayılmayan Gün": int(status_counts.get("HARIC", 0)),
            "Pazar Kesintisi": int(group["Pazar Durumu"].eq("Kesildi").sum()),
        })
    summary = pd.DataFrame(records).sort_values("Personel").reset_index(drop=True)

    detail_columns = [
        "sicilno", "Personel", "Firma", "Bölüm", "Pozisyon", "Görev", "Yaka", "Tarih",
        "Giriş", "Çıkış", "MS_h", "NM_h", "FM_h", "NM Güncel_h", "FM Güncel_h", "FM→NM_h",
        "IZS_h", "YIZS_h", "SGKIZS_h", "UCZIZS_h", "RM_h", "EM_h", "İzin Açıklama",
        "Durum Açıklaması", "Kaynak Kod", "Kod", "Pazar Durumu",
    ]
    detail = daily[detail_columns].rename(columns={"sicilno": "Sicil No"}).reset_index(drop=True)
    return ReportResult(detail, monthly, summary, weekly.reset_index(drop=True), quality, period_start, period_end)


def available_periods(df: pd.DataFrame) -> list[tuple[int, int]]:
    normalized = normalize_columns(df)
    if "mesaitarih" not in normalized.columns:
        return []
    dates = pd.to_datetime(normalized["mesaitarih"], dayfirst=True, errors="coerce").dropna()
    return sorted({(int(value.year), int(value.month)) for value in dates})


def _write_dataframe(ws, df: pd.DataFrame, title: str, freeze: str = "A2") -> None:
    ws.title = title
    header_fill = PatternFill("solid", fgColor="16324F")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for col_idx, column in enumerate(df.columns, 1):
        cell = ws.cell(1, col_idx, column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(df.itertuples(index=False, name=None), 2):
        for col_idx, value in enumerate(row, 1):
            if pd.isna(value):
                value = None
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = Border(bottom=thin)
            if isinstance(value, pd.Timestamp):
                cell.number_format = "dd.mm.yyyy"
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 34
    for idx, column in enumerate(df.columns, 1):
        sample = [len(str(column))] + [len(str(value)) for value in df[column].head(150) if not pd.isna(value)]
        ws.column_dimensions[get_column_letter(idx)].width = min(max(sample, default=8) + 2, 42)


def create_excel_report(result: ReportResult) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    monthly_ws = wb.create_sheet()
    _write_dataframe(monthly_ws, result.monthly, "Aylık Puantaj", "H2")
    for cell in monthly_ws[1][7:]:
        day_number = int(str(cell.value)[:2])
        if pd.Timestamp(result.period_start.year, result.period_start.month, day_number).dayofweek >= 5:
            cell.fill = PatternFill("solid", fgColor="8A5A00")
    for row in monthly_ws.iter_rows(min_row=2, min_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            if cell.value in EXCEL_CODE_COLORS:
                cell.fill = PatternFill("solid", fgColor=EXCEL_CODE_COLORS[cell.value])

    summary = result.summary.copy()
    weekly = result.weekly.copy()
    detail = result.daily.copy()
    hour_columns = ["Normal Çalışma", "Fazla Mesai", "FM→NM Aktarım"]
    for column in hour_columns:
        summary[column] = summary[column].map(format_hours)
    for column in ["Hafta İçi NM", "Hafta Sonu Ham FM", "FM→NM Aktarım", "Toplam NM", "Kalan FM"]:
        weekly[column] = weekly[column].map(format_hours)
    for column in [c for c in detail.columns if c.endswith("_h")]:
        detail[column] = detail[column].map(format_hours)
    detail["Tarih"] = pd.to_datetime(detail["Tarih"])

    for frame, name, freeze in [
        (summary, "Personel Özeti", "A2"),
        (weekly, "Haftalık Kontrol", "A2"),
        (detail, "Günlük Detay", "H2"),
        (result.quality, "Veri Kalitesi", "A2"),
        (pd.DataFrame(CODE_LEGEND, columns=["Kod", "Açıklama"]), "Kodlar", "A2"),
    ]:
        ws = wb.create_sheet()
        _write_dataframe(ws, frame, name, freeze)

    meta = wb.create_sheet("Rapor Bilgisi", 0)
    meta.append(["AYLIK PUANTAJ RAPORU"])
    meta.append(["Dönem", f"{result.period_start:%m.%Y}"])
    meta.append(["Personel", len(result.summary)])
    meta.append(["Kayıt", len(result.daily)])
    meta.append(["Üretim zamanı", datetime.now().strftime("%d.%m.%Y %H:%M")])
    meta.append([])
    meta.append(["Not", "Çalışma hücreleri NM+FM toplamını; özet sayfası haftalık 45 saat aktarımı sonrasını gösterir."])
    meta["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    meta["A1"].fill = PatternFill("solid", fgColor="16324F")
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 95

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
