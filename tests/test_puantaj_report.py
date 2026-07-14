from datetime import time
from io import BytesIO

import pandas as pd
from openpyxl import Workbook

from puantaj_report import build_report, create_excel_report, read_puantaj_file, time_to_hours


def sample_frame():
    rows = []
    for day in range(1, 8):
        rows.append({
            "sicilno": "00001",
            "Ad": "TEST",
            "Soyad": "PERSONEL",
            "mesaitarih": pd.Timestamp(2026, 6, day),
            "NM": time(9) if day <= 4 else time(0),
            "FM": time(3) if day == 6 else time(0),
            "MS": time(9) if day <= 5 else time(0),
            "EM": time(9) if day == 5 else time(0),
            "IZS": time(0),
            "YIZS": time(0),
            "SGKIZS": time(0),
            "UCZIZS": time(0),
            "RM": time(0),
            "İzin Açıklama": "#__#",
            "Bölüm": "Üretim",
        })
    return pd.DataFrame(rows)


def test_time_to_hours_supports_excel_values():
    assert time_to_hours(time(8, 30)) == 8.5
    assert time_to_hours(0.5) == 12
    assert time_to_hours("09:15:00") == 9.25


def test_report_applies_weekend_transfer_and_sunday_cut():
    result = build_report(sample_frame(), 2026, 6)
    assert result.weekly.loc[0, "FM→NM Aktarım"] == 3
    assert result.weekly.loc[0, "Kalan FM"] == 0
    assert result.weekly.loc[0, "Pazar Durumu"] == "Kesildi"
    assert result.monthly.loc[0, "05 Cu"] == "E"
    assert result.monthly.loc[0, "07 Pz"] == "Z"


def test_excel_report_is_created():
    data = create_excel_report(build_report(sample_frame(), 2026, 6))
    assert data.startswith(b"PK")
    assert len(data) > 5000


def test_numeric_file_name_does_not_break_reader(tmp_path):
    path = tmp_path / "sample.xlsx"
    sample_frame().to_excel(path, index=False)
    with path.open("rb") as source:
        result = read_puantaj_file(source, filename=123)
    assert len(result) == 7


def test_sakra_horizontal_workbook_is_detected():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2026 Puantaj"
    sheet.cell(2, 3, " ADI SOYADI")
    sheet.cell(2, 13, "Yıl")
    sheet.cell(2, 14, "Ay")
    sheet.cell(3, 3, "TEST PERSONEL")
    sheet.cell(3, 13, 2026)
    sheet.cell(3, 14, 7)
    sheet.cell(3, 27, 9)  # 1 Temmuz 2026 Çarşamba, ilk günlük blokta AA.
    sheet.cell(3, 37, 9)  # İlk haftanın normal çalışma sonucu.
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    source = read_puantaj_file(buffer, "sakra.xlsx")
    result = build_report(source, 2026, 7)

    assert len(source) == 1
    assert result.monthly.loc[0, "01 Ça"] == 9
